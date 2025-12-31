"""
Generate LaTeX tables and Excel spreadsheets from experimental results.

This script converts CSV results into:
1. Formatted LaTeX tables for academic papers
2. Excel files with formatted tables for easy viewing/editing
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Optional, cast
import os
# Safe numeric coercion helper
from src.utils.num_utils import safe_to_float
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    # Import `dataframe_to_rows` lazily inside the Excel export helper to avoid static import errors
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl not available. Excel export will be disabled.")
    print("   Install with: pip install openpyxl")


def generate_mnist_comparison_table(csv_path: str, output_path: Optional[str] = None, excel_path: Optional[str] = None):
    """
    Generate LaTeX table and Excel file for MNIST optimizer comparison.
    
    Args:
        csv_path: Path to statistical comparisons CSV
        output_path: Optional path to save LaTeX file
        excel_path: Optional path to save Excel file
    """
    df = pd.read_csv(csv_path)
    
    # Add significance stars column for display
    def _format_p(p):
        try:
            p_f = float(p)
        except Exception:
            p_f = float('nan')
        if p_f < 0.001:
            return f"{p_f:.4f}***"
        if p_f < 0.01:
            return f"{p_f:.4f}**"
        if p_f < 0.05:
            return f"{p_f:.4f}*"
        return f"{p_f:.4f}"

    df['p-value (formatted)'] = df['p-value'].apply(_format_p)
    
    # Format Cohen's d (safe)
    if "Cohen's d" in df.columns:
        def _format_effect_size(d):
            try:
                v = safe_to_float(d)
                if np.isnan(v):
                    return "---"
                return f"{v:.3f}"
            except Exception:
                return "---"
        df["Cohen's d (formatted)"] = df["Cohen's d"].apply(_format_effect_size)
    
    # Format power (safe)
    if 'Observed power' in df.columns:
        def _format_power(p):
            try:
                v = safe_to_float(p)
                if np.isnan(v):
                    return "---"
                return f"{v:.3f}"
            except Exception:
                return "---"
        df['Power (formatted)'] = df['Observed power'].apply(_format_power)
    
    # === LaTeX Table ===
    latex_lines = []
    latex_lines.append("\\begin{table}[htbp]")
    latex_lines.append("\\centering")
    latex_lines.append("\\caption{Statistical Comparison of Optimizers on MNIST (SimpleMLP, 10 epochs)}")
    latex_lines.append("\\label{tab:mnist_comparison}")
    latex_lines.append("\\begin{tabular}{llcccccc}")
    latex_lines.append("\\toprule")
    latex_lines.append("Optimizer A & Optimizer B & $n$ & Mean A & Mean B & $p$-value & Cohen's $d$ & Power \\\\")
    latex_lines.append("\\midrule")
    
    for _, row in df.iterrows():
        opt_a = str(row['Optimizer A']).replace('_', '\\_')
        opt_b = str(row['Optimizer B']).replace('_', '\\_')
        n = int(row['n'])
        mean_a = f"{row['Mean A']:.4f}"
        mean_b = f"{row['Mean B']:.4f}"
        
        # Format p-value with significance stars (coerce safely to float)
        p_val_f = safe_to_float(row.get('p-value', np.nan))
        if not np.isnan(p_val_f) and p_val_f < 0.001:
            p_str = f"{p_val_f:.4f}***"
        elif not np.isnan(p_val_f) and p_val_f < 0.01:
            p_str = f"{p_val_f:.4f}**"
        elif not np.isnan(p_val_f) and p_val_f < 0.05:
            p_str = f"{p_val_f:.4f}*"
        elif not np.isnan(p_val_f):
            p_str = f"{p_val_f:.4f}"
        else:
            p_str = "---"
        
        cohens_d = safe_to_float(row.get("Cohen's d", np.nan))
        cohens_str = f"{cohens_d:.3f}" if not np.isnan(cohens_d) else "---"

        power = safe_to_float(row.get('Observed power', np.nan))
        power_str = f"{power:.3f}" if not np.isnan(power) else "---"
        
        latex_lines.append(f"{opt_a} & {opt_b} & {n} & {mean_a} & {mean_b} & {p_str} & {cohens_str} & {power_str} \\\\")
    
    latex_lines.append("\\bottomrule")
    latex_lines.append("\\end{tabular}")
    latex_lines.append("\\begin{tablenotes}")
    latex_lines.append("\\small")
    latex_lines.append("\\item Note: $p$-values are corrected using Holm-Bonferroni method. ")
    latex_lines.append("Significance levels: * $p < 0.05$, ** $p < 0.01$, *** $p < 0.001$. ")
    latex_lines.append("Cohen's $d$: small $|d| < 0.5$, medium $0.5 \\leq |d| < 0.8$, large $|d| \\geq 0.8$.")
    latex_lines.append("\\end{tablenotes}")
    latex_lines.append("\\end{table}")
    
    latex_content = "\n".join(latex_lines)
    
    if output_path:
        with open(output_path, 'w') as f:
            f.write(latex_content)
        print(f"✅ LaTeX table saved to: {output_path}")
    
    # === Excel Export ===
    if excel_path and OPENPYXL_AVAILABLE:
        _export_to_excel_mnist(df, excel_path)
    
    return latex_content


from typing import Optional

def _export_to_excel_mnist(df: pd.DataFrame, excel_path: Optional[str]):
    if excel_path is None:
        raise ValueError("excel_path must be provided")
    """Export MNIST comparison to Excel with formatting."""
    try:
        import importlib
        df_mod = importlib.import_module('openpyxl.utils.dataframe')
        dataframe_to_rows = df_mod.dataframe_to_rows
    except Exception as e:
        raise RuntimeError('openpyxl is not available at runtime; ensure OPENPYXL_AVAILABLE is True') from e
    
    wb = Workbook()
    ws = wb.active
    ws.title = "MNIST Comparison"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Select columns to export
    export_cols = ['Optimizer A', 'Optimizer B', 'n', 'Mean A', 'Std A', 'Mean B', 'Std B',
                   'p-value', 'Significant (α=0.05)', "Cohen's d", 'Observed power', 'Required n (80%)']
    df_export = df[[col for col in export_cols if col in df.columns]].copy()
    
    # Round numeric columns
    for col in ['Mean A', 'Std A', 'Mean B', 'Std B', 'p-value', "Cohen's d", 'Observed power']:
        if col in df_export.columns:
            df_export[col] = df_export[col].round(4)
    
    # Add headers
    headers = list(df_export.columns)
    ws.append(headers)
    
    # Style header row (only if header row exists)
    header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if header_row is not None:
        for cell in header_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
    
    # Add data rows
    for row_data in dataframe_to_rows(df_export, index=False, header=False):
        ws.append(row_data)
    
    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = center_align
    
    # Highlight significant results (p < 0.05)
    sig_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    p_col_idx = headers.index('p-value') + 1 if 'p-value' in headers else None
    
    if p_col_idx is not None:
        for row_idx in range(2, ws.max_row + 1):
            try:
                p_val = ws.cell(row_idx, p_col_idx).value
            except Exception:
                # Defensive: skip rows we can't access
                continue
            if isinstance(p_val, (int, float)) and p_val < 0.05:
                # Use iter_rows to fetch the row safely (avoids optional subscript warning)
                row_cells = ()
                for r in ws.iter_rows(min_row=row_idx, max_row=row_idx):
                    row_cells = r
                    break
                for cell in row_cells:
                    if cell is None:
                        continue
                    cell.fill = sig_fill

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        try:
            first_cell = column[0]
        except Exception:
            continue
        column_letter = getattr(first_cell, 'column_letter', None)
        if column_letter is None:
            # MergedCell stubs may not expose column_letter; fall back to numeric column index
            try:
                from openpyxl.utils import get_column_letter
                column_letter = get_column_letter(getattr(first_cell, 'column', 1))
            except Exception:
                column_letter = 'A'
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                # Handle None or non-stringable cell values
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add Notes worksheet with interpretation and methodology
    notes = [
        [""],
        ["Methodology:"],
        ["- n=10 seeds per optimizer"],
        ["- Paired statistical tests (when seeds match)"],
        ["- Holm-Bonferroni correction for multiple comparisons"],
        ["- α = 0.05 significance level"],
        [""],
        ["Interpretation:"],
        ["- p-value < 0.05: Statistically significant difference"],
        ["- Cohen's d: Effect size (small <0.5, medium 0.5-0.8, large >0.8)"],
        ["- Power: Probability of detecting true effect (target: 0.80)"],
        ["- Required n: Sample size needed for 80% power"],
        [""],
        ["Color coding:"],
        ["- Green rows: Significant differences (p < 0.05)"],
    ]

    notes_ws = wb.create_sheet("Notes")
    for row in notes:
        notes_ws.append(row)

    assert excel_path is not None
    wb.save(cast(str, excel_path))
    print(f"✅ Excel file saved to: {excel_path}")



    # Export data
    df_export = df.copy()
    
    # Add headers
    headers = list(df_export.columns)
    ws.append(headers)
    
    # Style header (only if header row exists)
    header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if header_row is not None:
        for cell in header_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
    
    # Add data
    for row_data in dataframe_to_rows(df_export, index=False, header=False):
        ws.append(row_data)
    
    # Style data
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = center_align
    
    # Highlight converged optimizers
    conv_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    conv_col_idx = headers.index('Converged (loss<1e-3)') + 1 if 'Converged (loss<1e-3)' in headers else None

    if conv_col_idx is not None:
        for row_idx in range(2, ws.max_row + 1):
            converged = ws.cell(row_idx, conv_col_idx).value
            try:
                conv_bool = bool(converged)
            except Exception:
                conv_bool = False
            if conv_bool:
                # Use iter_rows to fetch the row safely (avoids optional subscript warning)
                row_cells = ()
                for r in ws.iter_rows(min_row=row_idx, max_row=row_idx):
                    row_cells = r
                    break
                for cell in row_cells:
                    if cell is None:
                        continue
                    cell.fill = conv_fill
    
    # Auto-adjust columns
    for column in ws.columns:
        if not column:
            continue
        max_length = 0
        first_cell = column[0]
        column_letter = getattr(first_cell, 'column_letter', None)
        if column_letter is None:
            try:
                from openpyxl.utils import get_column_letter
                column_letter = get_column_letter(getattr(first_cell, 'column', 1))
            except Exception:
                column_letter = 'A'
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                # Handle None or non-stringable cell values
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    assert excel_path is not None
    wb.save(cast(str, excel_path))
    print(f"✅ Excel file saved to: {excel_path}")


def _export_to_excel_robustness(dfs_dict: dict, excel_path: Optional[str]):
    if excel_path is None:
        raise ValueError("excel_path must be provided")
    """Export robustness analysis to Excel with formatting."""
    try:
        import importlib
        df_mod = importlib.import_module('openpyxl.utils.dataframe')
        dataframe_to_rows = df_mod.dataframe_to_rows
    except Exception as e:
        raise RuntimeError('openpyxl is not available at runtime; ensure OPENPYXL_AVAILABLE is True') from e
    
    wb = Workbook()
    # Remove default sheet if present (some Workbook implementations may set active to None)
    active_sheet = wb.active
    if active_sheet is not None:
        wb.remove(active_sheet)  # Remove default sheet
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create a sheet for each test function
    for func_name, df in dfs_dict.items():
        ws = wb.create_sheet(func_name)
        
        # Headers
        headers = list(df.columns)
        ws.append(headers)
        
        # Style header (only if header row exists)
        header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if header_row is not None:
            for cell in header_row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
        
        # Add data
        for row_data in dataframe_to_rows(df, index=False, header=False):
            ws.append(row_data)
        
        # Style data
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = center_align
        
        # Color code by success rate
        success_col_idx = headers.index('success_rate') + 1 if 'success_rate' in headers else None
        if success_col_idx is not None:
            for row_idx in range(2, ws.max_row + 1):
                success_rate = ws.cell(row_idx, success_col_idx).value
                if isinstance(success_rate, (int, float)):
                    if success_rate >= 0.8:
                        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                    elif success_rate >= 0.5:
                        fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
                    elif success_rate > 0:
                        fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
                    else:
                        fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
                    
                    # Use iter_rows to fetch the row safely
                    for r in ws.iter_rows(min_row=row_idx, max_row=row_idx):
                        for cell in r:
                            if cell is None:
                                continue
                            cell.fill = fill
                        break
        
        # Auto-adjust columns
        for column in ws.columns:
            if not column:
                continue
            max_length = 0
            first_cell = column[0]
            column_letter = getattr(first_cell, 'column_letter', None)
            if column_letter is None:
                try:
                    from openpyxl.utils import get_column_letter
                    column_letter = get_column_letter(getattr(first_cell, 'column', 1))
                except Exception:
                    column_letter = 'A'
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary", 0)
    summary_ws.append(["Initial Condition Robustness Analysis"])
    summary_ws.append([])
    summary_ws.append(["Test Function", "Best Optimizer", "Success Rate"])
    
    for func_name, df in dfs_dict.items():
        best_row = df.loc[df['success_rate'].idxmax()]
        summary_ws.append([func_name, best_row['optimizer'], f"{best_row['success_rate']:.2%}"])
    
    assert excel_path is not None
    wb.save(cast(str, excel_path))
    print(f"✅ Excel file saved to: {excel_path}")


def _export_to_excel_ablation(df: pd.DataFrame, excel_path: Optional[str]):
    """Export ablation study results to Excel with simple formatting."""
    if excel_path is None:
        raise ValueError("excel_path must be provided")
    try:
        import importlib
        df_mod = importlib.import_module('openpyxl.utils.dataframe')
        dataframe_to_rows = df_mod.dataframe_to_rows
    except Exception as e:
        raise RuntimeError('openpyxl is not available at runtime; ensure OPENPYXL_AVAILABLE is True') from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Ablation Study"

    # Basic styles (reuse names from file scope)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = list(df.columns)
    ws.append(headers)
    # Style header (only if header row exists)
    header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if header_row is not None:
        for cell in header_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

    for row_data in dataframe_to_rows(df, index=False, header=False):
        ws.append(row_data)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = center_align

    # Highlight converged column if present
    conv_idx = headers.index('Converged (loss<1e-3)') + 1 if 'Converged (loss<1e-3)' in headers else None
    conv_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    if conv_idx is not None:
        for row_idx in range(2, ws.max_row + 1):
            try:
                val = ws.cell(row_idx, conv_idx).value
            except Exception:
                continue
            try:
                if bool(val):
                    # Use iter_rows to fetch the row safely
                    for r in ws.iter_rows(min_row=row_idx, max_row=row_idx):
                        for cell in r:
                            if cell is None:
                                continue
                            cell.fill = conv_fill
                        break
            except Exception:
                continue

    # Auto-adjust widths
    for column in ws.columns:
        if not column:
            continue
        max_length = 0
        first_cell = column[0]
        column_letter = getattr(first_cell, 'column_letter', None)
        if column_letter is None:
            try:
                from openpyxl.utils import get_column_letter
                column_letter = get_column_letter(getattr(first_cell, 'column', 1))
            except Exception:
                column_letter = 'A'
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    assert excel_path is not None
    wb.save(cast(str, excel_path))
    print(f"✅ Excel file saved to: {excel_path}")


def generate_ablation_table(csv_path: str, output_path: Optional[str] = None, excel_path: Optional[str] = None):
    """
    Generate LaTeX table and optional Excel file for optimizer ablation study.
    """
    df = pd.read_csv(csv_path)
    
    latex_lines = []
    latex_lines.append("\\begin{table}[htbp]")
    latex_lines.append("\\centering")
    latex_lines.append("\\caption{Optimizer Ablation Study on Rosenbrock Function}")
    latex_lines.append("\\label{tab:ablation}")
    latex_lines.append("\\begin{tabular}{lcccc}")
    latex_lines.append("\\toprule")
    latex_lines.append("Optimizer & Final Loss & Min Loss & Iterations & Converged \\\\")
    latex_lines.append("\\midrule")
    
    for _, row in df.iterrows():
        opt = str(row.get('Optimizer', '')).replace('_', '\\_')
        final_loss = row['Final Loss']
        if np.isfinite(final_loss):
            final_str = f"{final_loss:.2e}"
        else:
            final_str = "DIV"
        
        min_loss = row['Min Loss']
        if np.isfinite(min_loss):
            min_str = f"{min_loss:.2e}"
        else:
            min_str = "---"
        
        iters = int(row['Iterations to Loss<1e-3'])
        if iters < 10000:
            iter_str = f"{iters}"
        else:
            iter_str = ">10k"
        
        # Ensure boolean check works for scalars/ndarrays/Series
        conv_val = row.get('Converged (loss<1e-3)', False)
        try:
            conv_bool = bool(np.asarray(conv_val))
        except Exception:
            conv_bool = bool(conv_val)
        converged = "\\checkmark" if conv_bool else "---"
        
        latex_lines.append(f"{opt} & {final_str} & {min_str} & {iter_str} & {converged} \\\\")
    
    latex_lines.append("\\bottomrule")
    latex_lines.append("\\end{tabular}")
    latex_lines.append("\\begin{tablenotes}")
    latex_lines.append("\\small")
    latex_lines.append("\\item Note: All optimizers started from $(-1.5, 2.0)$ with 10,000 maximum iterations. ")
    latex_lines.append("Convergence threshold: loss $< 10^{-3}$. DIV indicates divergence.")
    latex_lines.append("\\end{tablenotes}")
    latex_lines.append("\\end{table}")
    
    latex_content = "\n".join(latex_lines)
    
    if output_path:
        with open(output_path, 'w') as f:
            f.write(latex_content)
        print(f"✅ LaTeX table saved to: {output_path}")
    
    # === Excel Export ===
    if excel_path and OPENPYXL_AVAILABLE:
        _export_to_excel_ablation(df, excel_path)
    
    return latex_content


def generate_robustness_table(csv_paths: list, output_path: Optional[str] = None, excel_path: Optional[str] = None):
    """
    Generate LaTeX table and optional Excel file for initial condition robustness.
    """
    latex_lines = []
    latex_lines.append("\\begin{table}[htbp]")
    latex_lines.append("\\centering")
    latex_lines.append("\\caption{Initial Condition Robustness: Success Rate Across Test Functions}")
    latex_lines.append("\\label{tab:robustness}")
    
    # Read all CSVs
    dfs = {}
    for csv_path in csv_paths:
        # Extract function name from filename
        import os
        fname = os.path.basename(csv_path)
        if 'Rosenbrock' in fname:
            func_name = 'Rosenbrock'
        elif 'Quadratic' in fname or 'IllConditioned' in fname:
            func_name = 'Ill-Cond. Quad.'
        elif 'Saddle' in fname:
            func_name = 'Saddle Point'
        else:
            continue
        dfs[func_name] = pd.read_csv(csv_path)
    
    # Get unique optimizers
    all_opts = set()
    for df in dfs.values():
        all_opts.update(df['optimizer'].values)
    optimizers = sorted(all_opts)
    
    # Create table
    ncols = len(dfs) + 1
    latex_lines.append(f"\\begin{{tabular}}{{l{'c'*len(dfs)}}}")
    latex_lines.append("\\toprule")
    
    header = "Optimizer"
    for func_name in sorted(dfs.keys()):
        header += f" & {func_name}"
    header += " \\\\"
    latex_lines.append(header)
    latex_lines.append("\\midrule")
    
    for opt in optimizers:
        opt_clean = opt.replace('_', '\\_')
        row = f"{opt_clean}"
        for func_name in sorted(dfs.keys()):
            df = dfs[func_name]
            opt_row = df[df['optimizer'] == opt]
            if not opt_row.empty:
                success_rate = opt_row['success_rate'].values[0]
                row += f" & {success_rate:.2%}"
            else:
                row += " & ---"
        row += " \\\\"
        latex_lines.append(row)
    
    latex_lines.append("\\bottomrule")
    latex_lines.append("\\end{tabular}")
    latex_lines.append("\\begin{tablenotes}")
    latex_lines.append("\\small")
    latex_lines.append("\\item Note: Success rate = proportion of 30 initial conditions that converged to ")
    latex_lines.append("gradient norm $< 10^{-6}$ within 5,000 iterations.")
    latex_lines.append("\\end{tablenotes}")
    latex_lines.append("\\end{table}")
    
    latex_content = "\n".join(latex_lines)
    
    if output_path:
        with open(output_path, 'w') as f:
            f.write(latex_content)
        print(f"✅ LaTeX table saved to: {output_path}")
    
    # === Excel Export ===
    if excel_path and OPENPYXL_AVAILABLE:
        _export_to_excel_robustness(dfs, excel_path)
    
    return latex_content


def generate_summary_statistics(results_dir: str = 'results'):
    """
    Generate summary statistics for paper.
    """
    import glob
    
    print("\n" + "="*80)
    print("SUMMARY STATISTICS FOR PAPER")
    print("="*80)
    
    # MNIST results
    mnist_files = glob.glob(f"{results_dir}/NN_SimpleMLP_MNIST_*_final.csv")
    if mnist_files:
        print(f"\nMNIST Experiments:")
        print(f"   Total runs: {len(mnist_files)}")
        
        # Group by optimizer
        from collections import defaultdict
        by_optimizer = defaultdict(list)
        
        for f in mnist_files:
            df = pd.read_csv(f)
            eval_df = df[df['phase'] == 'eval']
            if not eval_df.empty:
                # Robust retrieval of final test accuracy: handle Series or ndarray and coerce to float
                try:
                    from src.utils.num_utils import safe_to_float
                    final_acc = safe_to_float(eval_df['test_accuracy'])
                    if np.isnan(final_acc):
                        # Could not extract a reliable final test accuracy; skip
                        continue
                except Exception:
                    # Could not extract final test accuracy reliably from this file; skip
                    continue
                # Extract optimizer name robustly (metadata JSON first, then parse)
                opt_name = 'Unknown'
                
                # Try metadata JSON first
                meta_path = f.replace('.csv', '_meta.json')
                if os.path.exists(meta_path):
                    try:
                        with open(meta_path, 'r') as mf:
                            import json
                            meta = json.load(mf)
                            opt_name = meta.get('optimizer', 'Unknown')
                    except (json.JSONDecodeError, IOError):
                        pass
                
                # Fallback to filename parsing
                if opt_name == 'Unknown':
                    basename = os.path.basename(f)
                    parts = basename.split('_')
                    if len(parts) >= 4:
                        opt_name = parts[3]
                
                if opt_name != 'Unknown':
                    by_optimizer[opt_name].append(final_acc)
        
        print(f"\n   Results by optimizer (Test Accuracy):")
        for opt, accs in sorted(by_optimizer.items()):
            mean_acc = np.mean(accs)
            std_acc = np.std(accs)
            print(f"   {opt:15s}: {mean_acc:.4f} ± {std_acc:.4f} (n={len(accs)})")
    
    # 2D experiments
    exp_files = glob.glob(f"{results_dir}/*-R-*.csv")  # Rosenbrock experiments
    if exp_files:
        print(f"\n📊 2D Optimization Experiments:")
        print(f"   Rosenbrock experiments: {len(exp_files)}")
    
    print("\n" + "="*80)


def main():
    """Generate all LaTeX tables."""
    results_dir = 'results'
    
    print("="*80)
    print("LATEX TABLE & EXCEL EXPORT GENERATION")
    print("="*80)
    
    # 1. MNIST comparison table
    mnist_stats = f"{results_dir}/mnist_statistical_comparisons_benchmark.csv"
    if Path(mnist_stats).exists():
        print("\n📄 Generating MNIST comparison table...")
        latex = generate_mnist_comparison_table(
            mnist_stats,
            f"{results_dir}/table_mnist_comparison.tex",
            f"{results_dir}/table_mnist_comparison.xlsx"
        )
    
    # 2. Ablation table
    ablation_csv = f"{results_dir}/optimizer_ablation_summary.csv"
    if Path(ablation_csv).exists():
        print("\n📄 Generating ablation table...")
        latex = generate_ablation_table(
            ablation_csv,
            f"{results_dir}/table_ablation.tex",
            f"{results_dir}/table_ablation.xlsx"
        )
    
    # 3. Robustness table
    robustness_files = [
        f"{results_dir}/initial_condition_robustness_summary_Rosenbrock.csv",
        f"{results_dir}/initial_condition_robustness_summary_IllConditionedQuadratic.csv",
        f"{results_dir}/initial_condition_robustness_summary_SaddlePoint.csv"
    ]
    existing_rob_files = [f for f in robustness_files if Path(f).exists()]
    if existing_rob_files:
        print("\n📄 Generating robustness table...")
        latex = generate_robustness_table(
            existing_rob_files,
            f"{results_dir}/table_robustness.tex",
            f"{results_dir}/table_robustness.xlsx"
        )
    
    # 4. Summary statistics
    generate_summary_statistics(results_dir)
    
    print("\n" + "="*80)
    print("✅ LaTeX tables and Excel files generated!")
    print("="*80)
    print(f"\nGenerated files in {results_dir}/:")
    print("   LaTeX tables:")
    print("      - table_mnist_comparison.tex")
    print("      - table_ablation.tex")
    print("      - table_robustness.tex")
    if OPENPYXL_AVAILABLE:
        print("   Excel files:")
        print("      - table_mnist_comparison.xlsx")
        print("      - table_ablation.xlsx")
        print("      - table_robustness.xlsx")
    else:
        print("\n⚠️  openpyxl not available - Excel export skipped")
        print("   Install with: pip install openpyxl")
    print("\nYou can include LaTeX tables in your document with:")
    print("   \\input{results/table_mnist_comparison.tex}")
    print("="*80)


if __name__ == '__main__':
    main()
